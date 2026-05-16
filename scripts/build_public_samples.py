#!/usr/bin/env python3
"""Generate minimal public sample files for docx/xlsx/hwpx (safe to commit)."""
from __future__ import annotations

import zipfile
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "samples"

W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
OD_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def write_docx_layout(path: Path) -> None:
    """DOCX with page metrics, spacing, and a floating anchor drawing."""
    document_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="{W_NS}"
  xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing"
  xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
  xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
  xmlns:wps="http://schemas.microsoft.com/office/word/2010/wordprocessingShape">
  <w:body>
    <w:p>
      <w:pPr><w:spacing w:before="240" w:after="120"/></w:pPr>
      <w:r><w:t>Body paragraph above float.</w:t></w:r>
    </w:p>
    <w:p>
      <w:r>
        <w:drawing>
          <wp:anchor distT="0" distB="0" distL="0" distR="0" simplePos="0"
                     relativeHeight="251658240" behindDoc="0" locked="0" layoutInCell="1" allowOverlap="1">
            <wp:simplePos x="0" y="0"/>
            <wp:positionH relativeFrom="page"><wp:posOffset>1600000</wp:posOffset></wp:positionH>
            <wp:positionV relativeFrom="page"><wp:posOffset>2400000</wp:posOffset></wp:positionV>
            <wp:extent cx="1200000" cy="600000"/>
            <wp:wrapSquare wrapText="bothSides"/>
            <a:graphic>
              <a:graphicData uri="http://schemas.microsoft.com/office/word/2010/wordprocessingShape">
                <wps:wsp>
                  <wps:txbx>
                    <w:txbxContent>
                      <w:p><w:r><w:t>Floating text box</w:t></w:r></w:p>
                    </w:txbxContent>
                  </wps:txbx>
                </wps:wsp>
              </a:graphicData>
            </a:graphic>
          </wp:anchor>
        </w:drawing>
      </w:r>
    </w:p>
    <w:sectPr>
      <w:pgSz w:w="12240" w:h="15840"/>
      <w:pgMar w:top="1440" w:right="1440" w:bottom="1440" w:left="1440"/>
    </w:sectPr>
  </w:body>
</w:document>"""
    content_types = f"""<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="{CT_NS}">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>"""
    rels = f"""<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="{REL_NS}">
  <Relationship Id="rId1" Type="{OD_REL}/officeDocument" Target="word/document.xml"/>
</Relationships>"""
    doc_rels = f"""<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="{REL_NS}"></Relationships>"""
    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels)
        zf.writestr("word/document.xml", document_xml)
        zf.writestr("word/_rels/document.xml.rels", doc_rels)


def write_docx(path: Path) -> None:
    document_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="{W_NS}">
  <w:body>
    <w:p>
      <w:pPr><w:pStyle w:val="Heading1"/></w:pPr>
      <w:r><w:t>Public DOCX Sample</w:t></w:r>
    </w:p>
    <w:p>
      <w:pPr><w:numPr><w:ilvl w:val="0"/></w:numPr></w:pPr>
      <w:r><w:t>List item one</w:t></w:r>
    </w:p>
    <w:tbl>
      <w:tr>
        <w:tc><w:p><w:r><w:t>Col A</w:t></w:r></w:p></w:tc>
        <w:tc><w:p><w:r><w:t>Col B</w:t></w:r></w:p></w:tc>
      </w:tr>
    </w:tbl>
  </w:body>
</w:document>"""
    content_types = f"""<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="{CT_NS}">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>"""
    rels = f"""<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="{REL_NS}">
  <Relationship Id="rId1" Type="{OD_REL}/officeDocument" Target="word/document.xml"/>
</Relationships>"""
    doc_rels = f"""<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="{REL_NS}"></Relationships>"""
    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels)
        zf.writestr("word/document.xml", document_xml)
        zf.writestr("word/_rels/document.xml.rels", doc_rels)


def write_hwpx(path: Path) -> None:
    section = """<?xml version="1.0" encoding="UTF-8"?>
<hs:sec xmlns:hs="http://www.hancom.co.kr/hwpml/2011/section"
        xmlns:hp="http://www.hancom.co.kr/hwpml/2011/paragraph">
  <hp:p><hp:run><hp:t>Public HWPX Sample</hp:t></hp:run></hp:p>
  <hp:tbl>
    <hp:tr>
      <hp:tc colSpan="2"><hp:p><hp:run><hp:t>Merged header</hp:t></hp:run></hp:p></hp:tc>
    </hp:tr>
    <hp:tr>
      <hp:tc><hp:p><hp:run><hp:t>A</hp:t></hp:run></hp:p></hp:tc>
      <hp:tc><hp:p><hp:run><hp:t>B</hp:t></hp:run></hp:p></hp:tc>
    </hp:tr>
  </hp:tbl>
</hs:sec>"""
    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr("Contents/section0.xml", section)


def write_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Demo"
    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    ws["A2"] = "accuracy"
    ws["B2"] = 0.95
    ws.merge_cells("A1:B1")
    wb.save(path)


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    write_docx(OUT / "openxml_parser_public_sample.docx")
    write_docx_layout(OUT / "openxml_parser_public_sample_layout.docx")
    write_xlsx(OUT / "openxml_parser_public_sample.xlsx")
    write_hwpx(OUT / "openxml_parser_public_sample.hwpx")
    print(f"Wrote samples under {OUT}")


if __name__ == "__main__":
    main()
