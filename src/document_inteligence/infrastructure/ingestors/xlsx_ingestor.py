from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from document_inteligence.domain.entities import (
    DocumentElement,
    DocumentPage,
    ElementType,
    ParsedDocument,
)
from document_inteligence.domain.repositories import DocumentIngestor
from document_inteligence.domain.value_objects import BBox
from document_inteligence.infrastructure.ingestors._ooxml_utils import clamp_01, synthetic_bbox


class XlsxIngestor(DocumentIngestor):
    def __init__(self, asset_output_dir: str | None = None) -> None:
        self._asset_output_dir = Path(asset_output_dir) if asset_output_dir else None

    def supports(self, path: str) -> bool:
        return Path(path).suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}

    def ingest(self, path: str) -> ParsedDocument:
        wb = load_workbook(path, data_only=True, read_only=False)
        pages: list[DocumentPage] = []

        for page_idx, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            elements = _sheet_to_elements(
                ws,
                page_number=page_idx,
                sheet_name=sheet_name,
                asset_output_dir=self._asset_output_dir,
            )
            pages.append(
                DocumentPage(
                    page_number=page_idx,
                    width=1.0,
                    height=1.0,
                    elements=elements,
                    metadata={
                        "source_format": "xlsx",
                        "sheet_name": sheet_name,
                        "layout_mode": "grid",
                    },
                )
            )
        wb.close()
        return ParsedDocument(source_path=str(path), pages=pages)


def _sheet_to_elements(
    ws,
    *,
    page_number: int,
    sheet_name: str,
    asset_output_dir: Path | None,
) -> list[DocumentElement]:
    if ws.max_row is None or ws.max_column is None:
        return []

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    merged_map = _build_merged_cell_map(ws)

    cells_meta: list[dict[str, object]] = []
    matrix_text: list[list[str]] = []

    for r in range(1, max_row + 1):
        row_vals: list[str] = []
        for c in range(1, max_col + 1):
            key = (r - 1, c - 1)
            info = merged_map.get(key)
            if info and info.get("is_spanned"):
                row_vals.append("")
                cells_meta.append(
                    {
                        "row": r - 1,
                        "col": c - 1,
                        "text": None,
                        "is_spanned": True,
                        "span_width": 1,
                        "span_height": 1,
                        "is_merge_origin": False,
                    }
                )
                continue

            value = ws.cell(row=r, column=c).value
            text = _cell_display(value)
            row_vals.append(text)

            span_w, span_h = 1, 1
            is_origin = True
            if info:
                span_w = int(info.get("span_width", 1))
                span_h = int(info.get("span_height", 1))
                is_origin = bool(info.get("is_origin", True))

            cells_meta.append(
                {
                    "row": r - 1,
                    "col": c - 1,
                    "text": text or None,
                    "is_spanned": False,
                    "span_width": span_w,
                    "span_height": span_h,
                    "is_merge_origin": is_origin,
                }
            )
        matrix_text.append(row_vals)

    pipe_lines = [" | ".join(row) for row in matrix_text if any(cell.strip() for cell in row)]
    pipe_text = "\n".join(pipe_lines) if pipe_lines else None

    col_widths = [1.0 for _ in range(max_col)]
    row_heights = [1.0 for _ in range(max_row)]

    elements: list[DocumentElement] = [
        DocumentElement(
            element_id=f"E_{page_number:03d}_0001",
            element_type=ElementType.TABLE,
            page_number=page_number,
            z_order=1,
            bbox=synthetic_bbox(0, row_height=0.9),
            text=pipe_text,
            metadata={
                "source_format": "xlsx",
                "sheet_name": sheet_name,
                "table_col_widths": col_widths,
                "table_row_heights": row_heights,
                "table_cells": cells_meta,
            },
        )
    ]

    order = 2
    for img_el in _extract_sheet_images(ws, page_number=page_number, asset_output_dir=asset_output_dir, start_order=order):
        elements.append(img_el)
        order += 1

    for chart_el in _extract_sheet_charts(ws, page_number=page_number, max_row=max_row, max_col=max_col, start_order=order):
        elements.append(chart_el)
        order += 1

    return elements


def _extract_sheet_images(
    ws,
    *,
    page_number: int,
    asset_output_dir: Path | None,
    start_order: int,
) -> list[DocumentElement]:
    images = getattr(ws, "_images", None) or []
    out: list[DocumentElement] = []
    order = start_order

    for idx, image in enumerate(images):
        anchor = getattr(image, "anchor", None)
        bbox = _anchor_bbox(anchor, ws, fallback_index=order)
        data = image._data()
        if not data:
            continue
        ext = "png"
        filename = f"sheet{page_number}_img{idx + 1}.{ext}"
        rel_path = filename
        if asset_output_dir is not None:
            asset_output_dir.mkdir(parents=True, exist_ok=True)
            out_path = asset_output_dir / filename
            out_path.write_bytes(data() if callable(data) else data)
            rel_path = f"{asset_output_dir.name}/{filename}"

        out.append(
            DocumentElement(
                element_id=f"E_{page_number:03d}_{order:04d}",
                element_type=ElementType.IMAGE,
                page_number=page_number,
                z_order=order,
                bbox=bbox,
                metadata={
                    "source_format": "xlsx",
                    "image": {
                        "filename": filename,
                        "relative_path": rel_path,
                        "content_type": f"image/{ext}",
                    },
                },
            )
        )
        order += 1
    return out


def _extract_sheet_charts(
    ws,
    *,
    page_number: int,
    max_row: int,
    max_col: int,
    start_order: int,
) -> list[DocumentElement]:
    charts = getattr(ws, "_charts", None) or []
    out: list[DocumentElement] = []
    order = start_order
    for idx, chart in enumerate(charts):
        anchor = getattr(chart, "anchor", None)
        bbox = _anchor_bbox(anchor, ws, fallback_index=order)
        title = getattr(chart, "title", None)
        text = str(title) if title else f"Chart {idx + 1}"
        out.append(
            DocumentElement(
                element_id=f"E_{page_number:03d}_{order:04d}",
                element_type=ElementType.CHART,
                page_number=page_number,
                z_order=order,
                bbox=bbox,
                text=text,
                metadata={"source_format": "xlsx", "chart_index": idx},
            )
        )
        order += 1
    return out


def _anchor_bbox(anchor, ws, *, fallback_index: int) -> BBox:
    if anchor is None or ws.max_column is None or ws.max_row is None:
        return synthetic_bbox(fallback_index)

    try:
        col = anchor._from.col + 1
        row = anchor._from.row + 1
        max_col = max(ws.max_column, 1)
        max_row = max(ws.max_row, 1)
        x = clamp_01((col - 1) / max_col)
        y = clamp_01((row - 1) / max_row)
        w = clamp_01(1.0 / max_col)
        h = clamp_01(1.0 / max_row)
        return BBox(x=x, y=y, width=max(w, 0.05), height=max(h, 0.05))
    except Exception:
        return synthetic_bbox(fallback_index)


def _cell_display(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _build_merged_cell_map(ws) -> dict[tuple[int, int], dict[str, object]]:
    out: dict[tuple[int, int], dict[str, object]] = {}
    for merged in ws.merged_cells.ranges:
        min_row, min_col = merged.min_row - 1, merged.min_col - 1
        max_row, max_col = merged.max_row - 1, merged.max_col - 1
        span_w = max_col - min_col + 1
        span_h = max_row - min_row + 1
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                is_origin = r == min_row and c == min_col
                out[(r, c)] = {
                    "is_spanned": not is_origin,
                    "is_origin": is_origin,
                    "span_width": span_w if is_origin else 1,
                    "span_height": span_h if is_origin else 1,
                }
    return out
