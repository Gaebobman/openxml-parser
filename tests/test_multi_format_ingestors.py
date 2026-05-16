from __future__ import annotations

import io
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from document_inteligence.domain.entities import ElementType
from document_inteligence.infrastructure.ingestors.docx_ingestor import DocxIngestor
from document_inteligence.infrastructure.ingestors.hwpx_ingestor import HwpxIngestor
from document_inteligence.infrastructure.ingestors.registry import build_ingestors
from document_inteligence.infrastructure.ingestors.xlsx_ingestor import XlsxIngestor


W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
OD_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _write_minimal_docx(path: Path) -> None:
    document_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="{W_NS}">
  <w:body>
    <w:p><w:r><w:t>Hello DOCX</w:t></w:r></w:p>
    <w:tbl>
      <w:tr>
        <w:tc><w:p><w:r><w:t>A1</w:t></w:r></w:p></w:tc>
        <w:tc><w:p><w:r><w:t>B1</w:t></w:r></w:p></w:tc>
      </w:tr>
    </w:tbl>
  </w:body>
</w:document>"""
    content_types = f"""<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="{CT_NS}">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>"""
    rels = f"""<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="{REL_NS}">
  <Relationship Id="rId1" Type="{OD_REL}/officeDocument" Target="word/document.xml"/>
</Relationships>"""
    doc_rels = f"""<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="{REL_NS}"></Relationships>"""

    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels)
        zf.writestr("word/document.xml", document_xml)
        zf.writestr("word/_rels/document.xml.rels", doc_rels)


def _write_minimal_hwpx(path: Path) -> None:
    section = """<?xml version="1.0" encoding="UTF-8"?>
<hs:sec xmlns:hs="http://www.hancom.co.kr/hwpml/2011/section"
        xmlns:hp="http://www.hancom.co.kr/hwpml/2011/paragraph">
  <hp:p><hp:run><hp:t>Hello HWPX</hp:t></hp:run></hp:p>
  <hp:tbl>
    <hp:tr>
      <hp:tc><hp:p><hp:run><hp:t>Cell</hp:t></hp:run></hp:p></hp:tc>
    </hp:tr>
  </hp:tbl>
</hs:sec>"""
    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr("Contents/section0.xml", section)


def test_docx_ingestor_parses_text_and_table(tmp_path: Path) -> None:
    docx = tmp_path / "sample.docx"
    _write_minimal_docx(docx)
    parsed = DocxIngestor().ingest(str(docx))
    assert len(parsed.pages) >= 1
    types = {e.element_type for page in parsed.pages for e in page.elements}
    assert ElementType.TEXT in types
    assert ElementType.TABLE in types
    texts = [e.text for page in parsed.pages for e in page.elements if e.text]
    assert any("Hello DOCX" in t for t in texts)


def test_xlsx_ingestor_parses_sheet_as_table(tmp_path: Path) -> None:
    xlsx = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "foo"
    ws["B2"] = "bar"
    wb.save(xlsx)
    parsed = XlsxIngestor().ingest(str(xlsx))
    assert len(parsed.pages) == 1
    assert parsed.pages[0].metadata.get("sheet_name") == "Sheet1"
    table = next(e for e in parsed.pages[0].elements if e.element_type == ElementType.TABLE)
    assert "foo" in (table.text or "")
    assert table.metadata.get("table_cells")


def test_hwpx_ingestor_parses_section(tmp_path: Path) -> None:
    hwpx = tmp_path / "sample.hwpx"
    _write_minimal_hwpx(hwpx)
    parsed = HwpxIngestor().ingest(str(hwpx))
    assert len(parsed.pages) >= 1
    types = {e.element_type for page in parsed.pages for e in page.elements}
    assert ElementType.TEXT in types
    assert ElementType.TABLE in types


def test_registry_supports_extensions() -> None:
    ingestors = build_ingestors()
    assert any(i.supports("a.pptx") for i in ingestors)
    assert any(i.supports("a.docx") for i in ingestors)
    assert any(i.supports("a.xlsx") for i in ingestors)
    assert any(i.supports("a.hwpx") for i in ingestors)
    assert not any(i.supports("a.pdf") for i in ingestors)
